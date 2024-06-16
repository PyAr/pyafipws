#!/usr/bin/python
# -*- coding: utf8 -*-
# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by the
# Free Software Foundation; either version 3, or (at your option) any later
# version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of MERCHANTIBILITY
# or FITNESS FOR A PARTICULAR PURPOSE.  See the GNU General Public License
# for more details.

"""Test para formato_csv"""

__author__ = "Mariano Reingart <reingart@gmail.com>"
__copyright__ = "Copyright (C) 2010-2019 Mariano Reingart"
__license__ = "GPL 3.0"

import os
import sys
import unittest
import csv
import pytest
import tempfile
from openpyxl import Workbook, load_workbook
from io import StringIO
from unittest.mock import patch

# Add the 'formatos' directory to the Python module search path
formatos_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'formatos'))
sys.path.insert(0, formatos_dir)

from pyafipws.formatos.formato_csv import leer, aplanar, desaplanar, escribir

@pytest.mark.dontusefix
class TestLeerFunction(unittest.TestCase):
    def test_leer_csv_file(self):
        """
        Test that the leer function can read a valid CSV file correctly.
        """
        expected_data = [
            ['1', '6', '4004', '526', '20170826', '80', '30500010912', 'PES', '1.000000', '889.82', '186.86', '8.89', '0.00', '0.00', '1085.57', '1', '', '', '', '61233038185853', '20110619', 'A', '', 'S', '', '', '', '', '', 'mariano@sistemasagiles.com.ar', '21601192', '6443', 'Exento', '82016336', '8001', '', '', 'P1675G', 'COD2', '8.89', '1076.68', '0', '1.0', '0', '0', '5', '1.00', '20205766', '110170P', 'Impuesto municipal matanza', '889.82', '186.86', '7', '0', 'PRUEBA ART', 'SEGUNDO ART', '', '', 'Cliente XXX', '', '99', '', '889.82', 'Patricia 1 - Cdad de Buenos Aires - 1405 - Capital Federal - Argentina', '1076.68', '0', '30 Dias', '0.00', '1', '17', '1', '1801', '30500010912', '1802', 'BNA']
        ]
        result = leer('datos/facturas.csv', delimiter=';')
        self.assertEqual(result, expected_data)


    def test_leer_csv_custom_delimiter(self):
        """
        Test that the leer function can read a CSV file with a custom delimiter.
        """
        sample_csv_data = "Column1|Column2|Column3\nValue1|Value2|Value3\nValue4|Value5|Value6"
        expected_data = [
            ['Value1', 'Value2', 'Value3'],
            ['Value4', 'Value5', 'Value6']
        ]

        with patch('builtins.open', return_value=StringIO(sample_csv_data)):
            result = leer('data/sample.csv', delimiter='|')
        self.assertEqual(result, expected_data)


    def test_leer_csv_with_header(self):
        """
        Test that the leer function can read a CSV file with a header row correctly.
        """
        sample_csv_data = "Column1,Column2,Column3\nValue1,Value2,Value3\nValue4,Value5,Value6"
        expected_data = [
            ['Value1', 'Value2', 'Value3'],
            ['Value4', 'Value5', 'Value6']
        ]

        with patch('builtins.open', return_value=StringIO(sample_csv_data)):
            result = leer('data/sample.csv')
        self.assertEqual(result, expected_data)
    
    def test_leer_csv_without_header(self):
        """
        Test that the leer function can read a CSV file without a header row correctly.
        """
        sample_csv_data = "Value1,Value2,Value3\nValue4,Value5,Value6"
        expected_data = [
            ['Value1', 'Value2', 'Value3'],
            ['Value4', 'Value5', 'Value6']
        ]

        with patch('builtins.open', return_value=StringIO(sample_csv_data)):
            result = leer('data/sample.csv', header=False)
        self.assertEqual(result, expected_data)


    def test_leer_csv_with_whitespace(self):
        """
        Test that the leer function can handle leading/trailing whitespace in CSV values correctly.
        """
        sample_csv_data = "Column1,Column2,Column3\n Value1 , Value2 , Value3 \n Value4 , Value5 , Value6 "
        expected_data = [
            ['Value1', 'Value2', 'Value3'],
            ['Value4', 'Value5', 'Value6']
        ]

        with patch('builtins.open', return_value=StringIO(sample_csv_data)):
            result = leer('data/sample.csv')
        self.assertEqual(result, expected_data)


    def test_leer_csv_with_non_string_values(self):
        """
        Test that the leer function can handle non-string values in a CSV file correctly.
        """
        sample_csv_data = "Column1,Column2,Column3\n1,2.5,True\n4,5.7,False"
        expected_data = [
            ['1', '2.5', 'True'],
            ['4', '5.7', 'False']
        ]

        with patch('builtins.open', return_value=StringIO(sample_csv_data)):
            result = leer('data/sample.csv')
        self.assertEqual(result, expected_data)


    def test_leer_csv_different_dialect(self):
        """
        Test that the leer function can handle a CSV file with a different dialect correctly.
        """
        sample_csv_data = "Column1;Column2;Column3\nValue1;Value2;Value3\nValue4;Value5;Value6"
        expected_data = [
            ['Value1', 'Value2', 'Value3'],
            ['Value4', 'Value5', 'Value6']
        ]

        with patch('builtins.open', return_value=StringIO(sample_csv_data)):
            result = leer('data/sample.csv', delimiter=';')
        self.assertEqual(result, expected_data)


    def test_leer_csv_empty_file(self):
        """
        Test that the leer function handles an empty CSV file correctly.
        """
        sample_csv_data = ""
        expected_data = []

        with patch('builtins.open', return_value=StringIO(sample_csv_data)):
            result = leer('data/sample.csv')
        self.assertEqual(result, expected_data)
      
        
    def test_leer_xlsx_file_with_header(self):
        """
        Test that the leer function can read an Excel file with a header row correctly.
        """
        expected_data = [
            ['Value1', 'Value2', 'Value3'],
            ['Value4', 'Value5', 'Value6']
        ]

        # Create a temporary Excel file for testing
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet['A1'] = 'Column1'
            worksheet['B1'] = 'Column2'
            worksheet['C1'] = 'Column3'
            worksheet['A2'] = 'Value1'
            worksheet['B2'] = 'Value2'
            worksheet['C2'] = 'Value3'
            worksheet['A3'] = 'Value4'
            worksheet['B3'] = 'Value5'
            worksheet['C3'] = 'Value6'
            workbook.save(temp_file.name)

            result = leer(temp_file.name)
        self.assertEqual(result, expected_data)

        # Clean up the temporary file
        os.unlink(temp_file.name)

    def test_leer_xlsx_file_without_header(self):
        """
        Test that the leer function can read an Excel file without a header row correctly.
        """
        expected_data = [
            ['Value1', 'Value2', 'Value3'],
            ['Value4', 'Value5', 'Value6']
        ]

        # Create a temporary Excel file for testing
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet['A1'] = 'Value1'
            worksheet['B1'] = 'Value2'
            worksheet['C1'] = 'Value3'
            worksheet['A2'] = 'Value4'
            worksheet['B2'] = 'Value5'
            worksheet['C2'] = 'Value6'
            workbook.save(temp_file.name)

            result = leer(temp_file.name, header=False)
        self.assertEqual(result, expected_data)

        # Clean up the temporary file
        os.unlink(temp_file.name)

@pytest.mark.dontusefix
class TestAplanarFunction(unittest.TestCase):
    def test_aplanar_single_record(self):
        """
        Test that the aplanar function correctly flattens a single record.
        """
        reg = {
            "id": 1,
            "tipo_cbte": 1,
            "punto_vta": 1,
            "cbte_nro": 1,
            "fecha_cbte": "2023-06-08",
            "tipo_doc": 80,
            "nro_doc": "20123456789",
            "imp_total": 126,
            "detalles": [
                {
                    "codigo": "P1",
                    "ds": "Producto 1",
                    "qty": 2,
                    "umed": 7,
                    "precio": 50,
                    "importe": 100,
                }
            ],
            "ivas": [
                {
                    "iva_id": 5,
                    "base_imp": 100,
                    "importe": 21,
                }
            ],
            "tributos": [
                {
                    "tributo_id": 1,
                    "base_imp": 100,
                    "desc": "Tributo 1",
                    "alic": 5,
                    "importe": 5,
                }
            ],
            "forma_pago": "Contado",
        }

        expected_data = [
            [
                'id', 'tipo_cbte', 'punto_vta', 'cbt_numero', 'fecha_cbte', 'tipo_doc', 'nro_doc', 'moneda_id', 'moneda_ctz',
                'imp_neto', 'imp_iva', 'imp_trib', 'imp_op_ex', 'imp_tot_conc', 'imp_total', 'concepto', 'fecha_venc_pago',
                'fecha_serv_desde', 'fecha_serv_hasta', 'cae', 'fecha_vto', 'resultado', 'motivo', 'reproceso', 'nombre',
                'domicilio', 'localidad', 'telefono', 'categoria', 'email', 'numero_cliente', 'numero_orden_compra',
                'condicion_frente_iva', 'numero_cotizacion', 'numero_remito', 'obs_generales', 'obs_comerciales',
                'forma_pago', 'pdf', 'bonif1', 'cantidad1', 'cbte_nro', 'codigo1', 'cuit', 'dato_a1', 'dato_b1', 'dato_c1',
                'dato_d1', 'dato_e1', 'descripcion1', 'detalles', 'domicilio_cliente', 'id_impositivo', 'idioma',
                'idioma_cbte', 'imp_iva1', 'importe1', 'incoterms', 'incoterms_ds', 'iva_base_imp_1', 'iva_id1', 'iva_id_1',
                'iva_importe_1', 'ivas', 'localidad_cliente', 'nombre_cliente', 'numero_despacho1', 'pais_dst_cmp',
                'permiso_existente', 'precio1', 'provincia_cliente', 'telefono_cliente', 'tipo_expo', 'tributo_alic_1',
                'tributo_base_imp_1', 'tributo_desc_1', 'tributo_id_1', 'tributo_importe_1', 'tributos', 'umed1'
            ],
            [
                1, 1, 1, 1, '2023-06-08', 80, '20123456789', None, None, None, None, None, None, None, 126, None, None,
                None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
                None, None, 'Contado', '', None, 2, 1, 'P1', None, None, None, None, None, None, 'Producto 1',
                [{'codigo': 'P1', 'ds': 'Producto 1', 'qty': 2, 'umed': 7, 'precio': 50, 'importe': 100}], None, None,
                None, None, None, 100, None, None, 100, None, 5, 21,
                [{'iva_id': 5, 'base_imp': 100, 'importe': 21}], None, None, None, None, None, 50, None, None, None, 5,
                100, 'Tributo 1', 1, 5,
                [{'tributo_id': 1, 'base_imp': 100, 'desc': 'Tributo 1', 'alic': 5, 'importe': 5}], 7
            ]
        ]

        result = aplanar([reg])
        self.assertEqual(result, expected_data)


    def test_aplanar_multiple_records(self):
        """
        Test that the aplanar function correctly flattens multiple records.
        """
        regs = [
            {
                "id": 1,
                "tipo_cbte": 1,
                "punto_vta": 1,
                "cbte_nro": 1,
                "fecha_cbte": "2023-06-08",
                "tipo_doc": 80,
                "nro_doc": "20123456789",
                "imp_total": 126,
                "detalles": [
                    {
                        "codigo": "P1",
                        "ds": "Producto 1",
                        "qty": 2,
                        "umed": 7,
                        "precio": 50,
                        "importe": 100,
                    }
                ],
                "ivas": [
                    {
                        "iva_id": 5,
                        "base_imp": 100,
                        "importe": 21,
                    }
                ],
                "tributos": [
                    {
                        "tributo_id": 1,
                        "base_imp": 100,
                        "desc": "Tributo 1",
                        "alic": 5,
                        "importe": 5,
                    }
                ],
                "forma_pago": "Contado",
            },
            {
                "id": 2,
                "tipo_cbte": 6,
                "punto_vta": 1,
                "cbte_nro": 2,
                "fecha_cbte": "2023-06-09",
                "tipo_doc": 80,
                "nro_doc": "20987654321",
                "imp_total": 200,
                "detalles": [
                    {
                        "codigo": "P2",
                        "ds": "Producto 2",
                        "qty": 1,
                        "umed": 7,
                        "precio": 200,
                        "importe": 200,
                    }
                ],
                "ivas": [
                    {
                        "iva_id": 5,
                        "base_imp": 200,
                        "importe": 42,
                    }
                ],
                "tributos": [],
                "forma_pago": "Tarjeta de Crédito",
            },
        ]

        expected_data = [
            [
                'id', 'tipo_cbte', 'punto_vta', 'cbt_numero', 'fecha_cbte', 'tipo_doc', 'nro_doc', 'moneda_id', 'moneda_ctz',
                'imp_neto', 'imp_iva', 'imp_trib', 'imp_op_ex', 'imp_tot_conc', 'imp_total', 'concepto', 'fecha_venc_pago',
                'fecha_serv_desde', 'fecha_serv_hasta', 'cae', 'fecha_vto', 'resultado', 'motivo', 'reproceso', 'nombre',
                'domicilio', 'localidad', 'telefono', 'categoria', 'email', 'numero_cliente', 'numero_orden_compra',
                'condicion_frente_iva', 'numero_cotizacion', 'numero_remito', 'obs_generales', 'obs_comerciales',
                'forma_pago', 'pdf', 'bonif1', 'cantidad1', 'cbte_nro', 'codigo1', 'cuit', 'dato_a1', 'dato_b1', 'dato_c1',
                'dato_d1', 'dato_e1', 'descripcion1', 'detalles', 'domicilio_cliente', 'id_impositivo', 'idioma',
                'idioma_cbte', 'imp_iva1', 'importe1', 'incoterms', 'incoterms_ds', 'iva_base_imp_1', 'iva_id1', 'iva_id_1',
                'iva_importe_1', 'ivas', 'localidad_cliente', 'nombre_cliente', 'numero_despacho1', 'pais_dst_cmp',
                'permiso_existente', 'precio1', 'provincia_cliente', 'telefono_cliente', 'tipo_expo', 'tributo_alic_1',
                'tributo_base_imp_1', 'tributo_desc_1', 'tributo_id_1', 'tributo_importe_1', 'tributos', 'umed1'
            ],
            [
                1, 1, 1, 1, '2023-06-08', 80, '20123456789', None, None, None, None, None, None, None, 126, None, None,
                None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
                None, None, 'Contado', '', None, 2, 1, 'P1', None, None, None, None, None, None, 'Producto 1',
                [{'codigo': 'P1', 'ds': 'Producto 1', 'qty': 2, 'umed': 7, 'precio': 50, 'importe': 100}], None, None,
                None, None, None, 100, None, None, 100, None, 5, 21,
                [{'iva_id': 5, 'base_imp': 100, 'importe': 21}], None, None, None, None, None, 50, None, None, None, 5,
                100, 'Tributo 1', 1, 5,
                [{'tributo_id': 1, 'base_imp': 100, 'desc': 'Tributo 1', 'alic': 5, 'importe': 5}], 7
            ],
            [
                2, 6, 1, 2, '2023-06-09', 80, '20987654321', None, None, None, None, None, None, None, 200, None, None,
                None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
                None, None, 'Tarjeta de Crédito', '', None, 1, 2, 'P2', None, None, None, None, None, None, 'Producto 2',
                [{'codigo': 'P2', 'ds': 'Producto 2', 'qty': 1, 'umed': 7, 'precio': 200, 'importe': 200}], None, None,
                None, None, None, 200, None, None, 200, None, 5, 42,
                [{'iva_id': 5, 'base_imp': 200, 'importe': 42}], None, None, None, None, None, 200, None, None, None,
                None, None, None, None, None, [], 7
            ]
        ]

        result = aplanar(regs)
        self.assertEqual(result, expected_data)


@pytest.mark.dontusefix
class TestDesplanarFunction(unittest.TestCase):
    def test_desaplanar_single_record(self):
        """
        Test that the desaplanar function correctly converts a single record from flattened format to structured format.
        """
        filas = [
            ['id', 'tipo_cbte', 'punto_vta', 'cbt_numero', 'fecha_cbte', 'tipo_doc', 'nro_doc', 'moneda_id', 'moneda_ctz',
            'imp_neto', 'imp_iva', 'imp_trib', 'imp_op_ex', 'imp_tot_conc', 'imp_total', 'concepto', 'fecha_venc_pago',
            'fecha_serv_desde', 'fecha_serv_hasta', 'cae', 'fecha_vto', 'resultado', 'motivo', 'reproceso', 'nombre',
            'domicilio', 'localidad', 'telefono', 'categoria', 'email', 'numero_cliente', 'numero_orden_compra',
            'condicion_frente_iva', 'numero_cotizacion', 'numero_remito', 'obs_generales', 'obs_comerciales',
            'forma_pago', 'pdf', 'codigo1', 'descripcion1', 'cantidad1', 'umed1', 'precio1', 'importe1', 'bonif1',
            'iva_id1', 'imp_iva1', 'tributo_id_1', 'tributo_desc_1', 'tributo_base_imp_1', 'tributo_alic_1',
            'tributo_importe_1', 'iva_id_1', 'iva_base_imp_1', 'iva_importe_1'],
            [1, 1, 1, 1, '2023-06-08', 80, '20123456789', 'PES', 1, 100, 21, 5, 0, 0, 126, 1, '2023-06-08', '2023-06-08',
            '2023-06-08', '1234567890', '2023-06-18', 'A', '', '', 'John Doe', '123 Main St', 'City', '1234567890', 'A',
            'john@example.com', 'ABC123', 'OC123', 'Responsable Inscripto', 'COT123', 'REM123', 'Observaciones generales',
            'Observaciones comerciales', 'Contado', '', 'P1', 'Producto 1', 2, 7, 50, 100, 0, 5, 21, 1, 'Tributo 1', 100,
            5, 5, 5, 100, 21]
        ]

        expected_data = [
            {
                'cbte_nro': 1,
                'tipo_cbte': 1,
                'punto_vta': 1,
                'cbt_numero': 1,
                'fecha_cbte': '2023-06-08',
                'concepto': 1,
                'moneda_id': 'PES',
                'moneda_ctz': 1,
                'tipo_doc': 80,
                'nro_doc': '20123456789',
                'email': 'john@example.com',
                'numero_cliente': 'ABC123',
                'numero_orden_compra': 'OC123',
                'condicion_frente_iva': 'Responsable Inscripto',
                'numero_cotizacion': 'COT123',
                'numero_remito': 'REM123',
                'imp_total': 126,
                'imp_tot_conc': 0,
                'imp_neto': 100,
                'imp_iva': 21,
                'imp_trib': 5,
                'imp_op_ex': 0,
                'fecha_serv_desde': '2023-06-08',
                'fecha_serv_hasta': '2023-06-08',
                'fecha_venc_pago': '2023-06-08',
                'obs_generales': 'Observaciones generales',
                'obs_comerciales': 'Observaciones comerciales',
                'resultado': 'A',
                'cae': '1234567890',
                'fecha_vto': '2023-06-18',
                'reproceso': '',
                'motivo': '',
                'id': 1,
                'detalles': [
                    {
                        'codigo': 'P1',
                        'ds': 'Producto 1',
                        'umed': 7,
                        'qty': 2,
                        'precio': 50,
                        'importe': 100,
                        'iva_id': 5,
                        'imp_iva': 21,
                        'bonif': None,
                        'despacho': False,
                        'dato_a': False,
                        'dato_b': False,
                        'dato_c': False,
                        'dato_d': False,
                        'dato_e': False
                    }
                ],
                'tributos': [
                    {
                        'tributo_id': 1,
                        'desc': 'Tributo 1',
                        'base_imp': 100,
                        'alic': 5,
                        'importe': 5
                    }
                ],
                'ivas': [
                    {
                        'iva_id': 5,
                        'base_imp': 100,
                        'importe': 21
                    }
                ],
                'permisos': [],
                'opcionales': [],
                'cbtes_asoc': [],
                'forma_pago': 'Contado',
                'datos': [
                    {'campo': 'nombre', 'valor': 'John Doe', 'pagina': ''},
                    {'campo': 'domicilio', 'valor': '123 Main St', 'pagina': ''},
                    {'campo': 'localidad', 'valor': 'City', 'pagina': ''},
                    {'campo': 'telefono', 'valor': '1234567890', 'pagina': ''},
                    {'campo': 'categoria', 'valor': 'A', 'pagina': ''},
                    {'campo': 'pdf', 'valor': '', 'pagina': ''}
                ]
            }
        ]

        result = desaplanar(filas)
        self.assertEqual(result, expected_data)

    
    def test_desaplanar_multiple_records(self):
        """
        Test that the desaplanar function correctly converts multiple records from flattened format to structured format.
        """
        filas = [
            ['id', 'tipo_cbte', 'punto_vta', 'cbt_numero', 'fecha_cbte', 'tipo_doc', 'nro_doc', 'moneda_id', 'moneda_ctz',
            'imp_neto', 'imp_iva', 'imp_trib', 'imp_op_ex', 'imp_tot_conc', 'imp_total', 'concepto', 'fecha_venc_pago',
            'fecha_serv_desde', 'fecha_serv_hasta', 'cae', 'fecha_vto', 'resultado', 'motivo', 'reproceso', 'nombre',
            'domicilio', 'localidad', 'telefono', 'categoria', 'email', 'numero_cliente', 'numero_orden_compra',
            'condicion_frente_iva', 'numero_cotizacion', 'numero_remito', 'obs_generales', 'obs_comerciales',
            'forma_pago', 'pdf', 'codigo1', 'descripcion1', 'cantidad1', 'umed1', 'precio1', 'importe1', 'bonif1',
            'iva_id1', 'imp_iva1', 'tributo_id_1', 'tributo_desc_1', 'tributo_base_imp_1', 'tributo_alic_1',
            'tributo_importe_1', 'iva_id_1', 'iva_base_imp_1', 'iva_importe_1'],
            [1, 1, 1, 1, '2023-06-08', 80, '20123456789', 'PES', 1, 100, 21, 5, 0, 0, 126, 1, '2023-06-08', '2023-06-08',
            '2023-06-08', '1234567890', '2023-06-18', 'A', '', '', 'John Doe', '123 Main St', 'City', '1234567890', 'A',
            'john@example.com', 'ABC123', 'OC123', 'Responsable Inscripto', 'COT123', 'REM123', 'Observaciones generales',
            'Observaciones comerciales', 'Contado', '', 'P1', 'Producto 1', 2, 7, 50, 100, 0, 5, 21, 1, 'Tributo 1', 100,
            5, 5, 5, 100, 21],
            [2, 1, 1, 2, '2023-06-09', 80, '20987654321', 'PES', 1, 200, 42, 10, 0, 0, 252, 1, '2023-06-09', '2023-06-09',
            '2023-06-09', '0987654321', '2023-06-19', 'A', '', '', 'Jane Smith', '456 Elm St', 'Town', '0987654321', 'B',
            'jane@example.com', 'XYZ789', 'OC456', 'Responsable Inscripto', 'COT456', 'REM456', 'Observaciones generales',
            'Observaciones comerciales', 'Tarjeta de Crédito', '', 'P2', 'Producto 2', 1, 7, 200, 200, 0, 5, 42, 2,
            'Tributo 2', 200, 5, 10, 5, 200, 42]
        ]

        expected_data = [
            {
                'cbte_nro': 1,
                'tipo_cbte': 1,
                'punto_vta': 1,
                'cbt_numero': 1,
                'fecha_cbte': '2023-06-08',
                'concepto': 1,
                'moneda_id': 'PES',
                'moneda_ctz': 1,
                'tipo_doc': 80,
                'nro_doc': '20123456789',
                'email': 'john@example.com',
                'numero_cliente': 'ABC123',
                'numero_orden_compra': 'OC123',
                'condicion_frente_iva': 'Responsable Inscripto',
                'numero_cotizacion': 'COT123',
                'numero_remito': 'REM123',
                'imp_total': 126,
                'imp_tot_conc': 0,
                'imp_neto': 100,
                'imp_iva': 21,
                'imp_trib': 5,
                'imp_op_ex': 0,
                'fecha_serv_desde': '2023-06-08',
                'fecha_serv_hasta': '2023-06-08',
                'fecha_venc_pago': '2023-06-08',
                'obs_generales': 'Observaciones generales',
                'obs_comerciales': 'Observaciones comerciales',
                'resultado': 'A',
                'cae': '1234567890',
                'fecha_vto': '2023-06-18',
                'reproceso': '',
                'motivo': '',
                'id': 1,
                'detalles': [
                    {
                        'codigo': 'P1',
                        'ds': 'Producto 1',
                        'umed': 7,
                        'qty': 2,
                        'precio': 50,
                        'importe': 100,
                        'iva_id': 5,
                        'imp_iva': 21,
                        'bonif': None,
                        'despacho': False,
                        'dato_a': False,
                        'dato_b': False,
                        'dato_c': False,
                        'dato_d': False,
                        'dato_e': False
                    }
                ],
                'tributos': [
                    {
                        'tributo_id': 1,
                        'desc': 'Tributo 1',
                        'base_imp': 100,
                        'alic': 5,
                        'importe': 5
                    }
                ],
                'ivas': [
                    {
                        'iva_id': 5,
                        'base_imp': 100,
                        'importe': 21
                    }
                ],
                'permisos': [],
                'opcionales': [],
                'cbtes_asoc': [],
                'forma_pago': 'Contado',
                'datos': [
                    {'campo': 'nombre', 'valor': 'John Doe', 'pagina': ''},
                    {'campo': 'domicilio', 'valor': '123 Main St', 'pagina': ''},
                    {'campo': 'localidad', 'valor': 'City', 'pagina': ''},
                    {'campo': 'telefono', 'valor': '1234567890', 'pagina': ''},
                    {'campo': 'categoria', 'valor': 'A', 'pagina': ''},
                    {'campo': 'pdf', 'valor': '', 'pagina': ''}
                ]
            },
            {
                'cbte_nro': 2,
                'tipo_cbte': 1,
                'punto_vta': 1,
                'cbt_numero': 2,
                'fecha_cbte': '2023-06-09',
                'concepto': 1,
                'moneda_id': 'PES',
                'moneda_ctz': 1,
                'tipo_doc': 80,
                'nro_doc': '20987654321',
                'email': 'jane@example.com',
                'numero_cliente': 'XYZ789',
                'numero_orden_compra': 'OC456',
                'condicion_frente_iva': 'Responsable Inscripto',
                'numero_cotizacion': 'COT456',
                'numero_remito': 'REM456',
                'imp_total': 252,
                'imp_tot_conc': 0,
                'imp_neto': 200,
                'imp_iva': 42,
                'imp_trib': 10,
                'imp_op_ex': 0,
                'fecha_serv_desde': '2023-06-09',
                'fecha_serv_hasta': '2023-06-09',
                'fecha_venc_pago': '2023-06-09',
                'obs_generales': 'Observaciones generales',
                'obs_comerciales': 'Observaciones comerciales',
                'resultado': 'A',
                'cae': '0987654321',
                'fecha_vto': '2023-06-19',
                'reproceso': '',
                'motivo': '',
                'id': 2,
                'detalles': [
                    {
                        'codigo': 'P2',
                        'ds': 'Producto 2',
                        'umed': 7,
                        'qty': 1,
                        'precio': 200,
                        'importe': 200,
                        'iva_id': 5,
                        'imp_iva': 42,
                        'bonif': None,
                        'despacho': False,
                        'dato_a': False,
                        'dato_b': False,
                        'dato_c': False,
                        'dato_d': False,
                        'dato_e': False
                    }
                ],
                'tributos': [
                    {
                        'tributo_id': 2,
                        'desc': 'Tributo 2',
                        'base_imp': 200,
                        'alic': 5,
                        'importe': 10
                    }
                ],
                'ivas': [
                    {
                        'iva_id': 5,
                        'base_imp': 200,
                        'importe': 42
                    }
                ],
                'permisos': [],
                'opcionales': [],
                'cbtes_asoc': [],
                'forma_pago': 'Tarjeta de Crédito',
                'datos': [
                    {'campo': 'nombre', 'valor': 'Jane Smith', 'pagina': ''},
                    {'campo': 'domicilio', 'valor': '456 Elm St', 'pagina': ''},
                    {'campo': 'localidad', 'valor': 'Town', 'pagina': ''},
                    {'campo': 'telefono', 'valor': '0987654321', 'pagina': ''},
                    {'campo': 'categoria', 'valor': 'B', 'pagina': ''},
                    {'campo': 'pdf', 'valor': '', 'pagina': ''}
                ]
            }
        ]

        result = desaplanar(filas)
        self.assertEqual(result, expected_data)


@pytest.mark.dontusefix
class TestEscribirFunction(unittest.TestCase):
    def test_escribir_facturas_csv(self):
        """
        Test that the escribir function can write data from facturas.csv correctly.
        """
        # Read the contents of facturas.csv
        with open('datos/facturas.csv', 'r', newline='') as file:
            csv_reader = csv.reader(file, delimiter=';')
            filas = [row for row in csv_reader]

        # Write the data to a new file using the escribir function
        filename = 'test_facturas_output.csv'
        escribir(filas, filename)

        # Check if the file was created
        self.assertTrue(os.path.exists(filename))

        # Read the contents of the output file and compare with the original data
        with open(filename, 'r', newline='') as file:
            csv_reader = csv.reader(file, delimiter=';')
            result = [row for row in csv_reader]
        self.assertEqual(result, filas)

        # Clean up the test file
        os.remove(filename)

    def test_escribir_facturas_xlsx(self):
        """
        Test that the escribir function can write data from facturas.csv to an XLSX file correctly.
        """
        # Read the contents of facturas.csv
        with open('datos/facturas.csv', 'r', newline='') as file:
            csv_reader = csv.reader(file, delimiter=';')
            filas = [row for row in csv_reader]

        # Write the data to an XLSX file using the escribir function
        filename = 'test_facturas_output.xlsx'
        escribir(filas, filename)

        # Check if the file was created
        self.assertTrue(os.path.exists(filename))

        # Read the contents of the XLSX file and compare with the original data
        workbook = load_workbook(filename)
        sheet = workbook.active
        result = []
        for row in sheet.iter_rows(values_only=True):
            # Convert None values to empty strings
            row = ['' if cell is None else str(cell) for cell in row]
            result.append(row)
        self.assertEqual(result, filas)

        # Clean up the test file
        os.remove(filename)
  
if __name__ == '__main__':
    unittest.main()
