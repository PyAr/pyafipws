#!/usr/bin/python
# -*- coding: utf8 -*-
# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published by the
# Free Software Foundation; either version 3, or (at your option) any later
# version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of MERCHANTIBILITY
# or FITNESS FOR A PARTICULAR PURPOSE.  See the GNU Lesser General Public License
# for more details.

"Módulo para manejo de archivos CSV (planillas de cálculo)"
from __future__ import print_function

from builtins import zip
from builtins import range
from past.builtins import basestring

__author__ = "Mariano Reingart (reingart@gmail.com)"
__copyright__ = "Copyright (C) 2010 Mariano Reingart"
__license__ = "LGPL-3.0-or-later"

import csv
from decimal import Decimal
import os
from openpyxl import load_workbook

def leer(fn="entrada.csv", delimiter=";", header=True):
    "Analiza un archivo CSV y devuelve una lista de listas con los datos"
    ext = os.path.splitext(fn)[1].lower()
    items = []

    if ext == ".csv":
        with open(fn, "r", encoding="utf-8") as csvfile:
            try:
                dialect = csv.Sniffer().sniff(csvfile.read(256), delimiters=[delimiter, ","])
            except csv.Error:
                dialect = csv.excel
                dialect.delimiter = delimiter
            csvfile.seek(0)
            csv_reader = csv.reader(csvfile, dialect)
            if header:
                next(csv_reader, None)  # Skip the header row
            for row in csv_reader:
                items.append([c.strip() if isinstance(c, str) else c for c in row])

    elif ext == ".xlsx":
        wb = load_workbook(filename=fn)
        ws1 = wb.active
        rows = ws1.iter_rows(values_only=True)
        if header:
            next(rows, None)  # Skip the header row
        for row in rows:
            items.append([cell for cell in row])

    return items
    # TODO: return desaplanar(items)
    

def aplanar(regs):
    "Convierte una estructura python en planilla CSV (PyRece)"

    from formato_xml import MAP_ENC

    filas = []
    for reg in regs:
        fila = {}

        # recorrer campos obligatorios:
        for k in MAP_ENC:
            fila[k] = reg.get(k)

        fila["forma_pago"] = reg.get("forma_pago", "")
        fila["pdf"] = reg.get("pdf", "")

        # datos adicionales (escalares):
        for k, v in list(reg.items()):
            if k not in MAP_ENC and isinstance(k, (basestring, int)):
                fila[k] = v

        # por compatibilidad con pyrece:
        if reg.get("cbte_nro"):
            fila["cbt_numero"] = reg["cbte_nro"]

        for i, det in enumerate(reg.get("detalles", []), start=1):
            fila.update(
                {
                    "codigo%s" % i: det.get("codigo", ""),
                    "descripcion%s" % i: det.get("ds", ""),
                    "umed%s" % i: det.get("umed"),
                    "cantidad%s" % i: det.get("qty"),
                    "precio%s" % i: det.get("precio"),
                    "importe%s" % i: det.get("importe"),
                    "iva_id%s" % i: det.get("iva_id"),
                    "imp_iva%s" % i: det.get("imp_iva"),
                    "bonif%s" % i: det.get("bonif"),
                    "numero_despacho%s" % i: det.get("despacho"),
                    "dato_a%s" % i: det.get("dato_a"),
                    "dato_b%s" % i: det.get("dato_b"),
                    "dato_c%s" % i: det.get("dato_c"),
                    "dato_d%s" % i: det.get("dato_d"),
                    "dato_e%s" % i: det.get("dato_e"),
                }
            )

        for i, iva in enumerate(reg.get("ivas", []), start=1):
            fila.update(
                {
                    "iva_id_%s" % i: iva.get("iva_id"),
                    "iva_base_imp_%s" % i: iva.get("base_imp"),
                    "iva_importe_%s" % i: iva.get("importe"),
                }
            )

        for i, tributo in enumerate(reg.get("tributos", []), start=1):
            fila.update(
                {
                    "tributo_id_%s" % i: tributo.get("tributo_id"),
                    "tributo_base_imp_%s" % i: tributo.get("base_imp"),
                    "tributo_desc_%s" % i: tributo.get("desc"),
                    "tributo_alic_%s" % i: tributo.get("alic"),
                    "tributo_importe_%s" % i: tributo.get("importe"),
                }
            )

        for i, opcional in enumerate(reg.get("opcionales", []), start=1):
            fila.update(
                {
                    "opcional_id_%s" % i: opcional.get("opcional_id"),
                    "opcional_valor_%s" % i: opcional.get("valor"),
                }
            )

        for i, cbte_asoc in enumerate(reg.get("cbtes_asoc", []), start=1):
            fila.update(
                {
                    "cbte_asoc_tipo_%s" % i: cbte_asoc.get("cbte_tipo"),
                    "cbte_asoc_pto_vta_%s" % i: cbte_asoc.get("cbte_punto_vta"),
                    "cbte_asoc_nro_%s" % i: cbte_asoc.get("cbte_nro"),
                    "cbte_asoc_cuit_%s" % i: cbte_asoc.get("cbte_cuit"),
                    "cbte_asoc_fecha_%s" % i: cbte_asoc.get("cbte_fecha"),
                }
            )

        filas.append(fila)

    cols = [
        "id",
        "tipo_cbte",
        "punto_vta",
        "cbt_numero",
        "fecha_cbte",
        "tipo_doc",
        "nro_doc",
        "moneda_id",
        "moneda_ctz",
        "imp_neto",
        "imp_iva",
        "imp_trib",
        "imp_op_ex",
        "imp_tot_conc",
        "imp_total",
        "concepto",
        "fecha_venc_pago",
        "fecha_serv_desde",
        "fecha_serv_hasta",
        "cae",
        "fecha_vto",
        "resultado",
        "motivo",
        "reproceso",
        "nombre",
        "domicilio",
        "localidad",
        "telefono",
        "categoria",
        "email",
        "numero_cliente",
        "numero_orden_compra",
        "condicion_frente_iva",
        "numero_cotizacion",
        "numero_remito",
        "obs_generales",
        "obs_comerciales",
        "forma_pago",
        "pdf",
    ]

    # filtro y ordeno las columnas
    l = [k for f in filas for k in list(f.keys())]
    s = set(l) - set(cols)
    cols.extend(sorted(s))

    ret = [cols]
    for fila in filas:
        ret.append([fila.get(k) for k in cols])

    return ret




def desaplanar(filas):
    "Dado una planilla, conviertir en estructura python"

    from formato_xml import MAP_ENC

    def max_li(colname):
        l = [int(k[len(colname) :]) + 1 for k in filas[0] if k.startswith(colname)]
        if l:
            tmp = max(l)
        if l and tmp:
            ##print "max_li(%s)=%s" % (colname, tmp)
            return tmp
        else:
            return 0

    regs = []
    for fila in filas[1:]:
        dic = dict([(filas[0][i], v) for i, v in enumerate(fila)])
        reg = {}

        # por compatibilidad con pyrece:
        reg["cbte_nro"] = dic["cbt_numero"]

        for k in MAP_ENC:
            if k in dic:
                reg[k] = dic.pop(k)

        reg["detalles"] = [
            {
                "codigo": ("codigo%s" % li) in dic and dic.pop("codigo%s" % li) or None,
                "ds": ("descripcion%s" % li) in dic
                and dic.pop("descripcion%s" % li)
                or None,
                "umed": ("umed%s" % li) in dic and dic.pop("umed%s" % li) or None,
                "qty": ("cantidad%s" % li) in dic
                and dic.pop("cantidad%s" % li)
                or None,
                "precio": ("precio%s" % li) in dic and dic.pop("precio%s" % li) or None,
                "importe": ("importe%s" % li) in dic
                and dic.pop("importe%s" % li)
                or None,
                "iva_id": ("iva_id%s" % li) in dic and dic.pop("iva_id%s" % li) or None,
                "imp_iva": ("imp_iva%s" % li) in dic
                and dic.pop("imp_iva%s" % li)
                or None,
                "bonif": ("bonif%s" % li) in dic and dic.pop("bonif%s" % li) or None,
                "despacho": ("numero_despacho%s" % li) in dic
                and dic.pop("numero_despacho%s" % li),
                "dato_a": ("dato_a%s" % li) in dic and dic.pop("dato_a%s" % li),
                "dato_b": ("dato_b%s" % li) in dic and dic.pop("dato_b%s" % li),
                "dato_c": ("dato_c%s" % li) in dic and dic.pop("dato_c%s" % li),
                "dato_d": ("dato_d%s" % li) in dic and dic.pop("dato_d%s" % li),
                "dato_e": ("dato_e%s" % li) in dic and dic.pop("dato_e%s" % li),
            }
            for li in range(1, max_li("cantidad"))
            if dic["cantidad%s" % li] is not None
        ]

        # descartar filas espurias vacias al final
        for det in reg["detalles"][::-1]:
            if any(det.values()):  # algun campo tiene dato termina
                break
            del reg["detalles"][-1]  # sino, borro último elemento

        reg["tributos"] = [
            {
                "tributo_id": dic.pop("tributo_id_%s" % li),
                "desc": dic.pop("tributo_desc_%s" % li),
                "base_imp": dic.pop("tributo_base_imp_%s" % li),
                "alic": dic.pop("tributo_alic_%s" % li),
                "importe": dic.pop("tributo_importe_%s" % li),
            }
            for li in range(1, max_li("tributo_id_"))
            if dic["tributo_id_%s" % li]
        ]

        reg["ivas"] = [
            {
                "iva_id": dic.pop("iva_id_%s" % li),
                "base_imp": dic.pop("iva_base_imp_%s" % li),
                "importe": dic.pop("iva_importe_%s" % li),
            }
            for li in range(1, max_li("iva_id_"))
            if dic["iva_id_%s" % li]
        ]

        reg["permisos"] = [
            {
                "id_permiso": dic.pop("id_permiso_%s" % li),
                "dst_merc": dic.pop("dst_merc_%s" % li),
            }
            for li in range(1, max_li("id_permiso_"))
            if dic["id_permiso_%s" % li]
        ]

        reg["opcionales"] = [
            {
                "opcional_id": dic.pop("opcional_id_%s" % li),
                "valor": dic.pop("opcional_valor_%s" % li),
            }
            for li in range(1, max_li("opcional_id_"))
            if dic["opcional_id_%s" % li]
        ]

        reg["cbtes_asoc"] = [
            {
                "cbte_tipo": dic.pop("cbte_asoc_tipo_%s" % li),
                "cbte_punto_vta": dic.pop("cbte_asoc_pto_vta_%s" % li),
                "cbte_nro": dic.pop("cbte_asoc_nro_%s" % li),
                "cbte_cuit": dic.pop("cbte_asoc_cuit_%s" % li),
                "cbte_fecha": dic.pop("cbte_asoc_fecha_%s" % li),
            }
            for li in range(1, max_li("cbte_asoc_tipo_"))
            if dic["cbte_asoc_tipo_%s" % li]
        ]

        reg["forma_pago"] = dic.pop("forma_pago")

        # agrego campos adicionales:
        reg["datos"] = [
            {
                "campo": campo,
                "valor": valor,
                "pagina": "",
            }
            for campo, valor in list(dic.items())
        ]

        regs.append(reg)

    return regs


def escribir(filas, fn="salida.csv", delimiter=";"):
    "Dado una lista de comprobantes (diccionarios), aplana y escribe"
    ext = os.path.splitext(fn)[1].lower()
    if ext == ".csv":
        with open(fn, "w", newline="") as f:
            csv_writer = csv.writer(f, dialect="excel", delimiter=delimiter)
            # TODO: filas = aplanar(regs)
            for fila in filas:
                csv_writer.writerow(fila)
    elif ext == ".xlsx":
        from openpyxl import Workbook

        wb = Workbook()
        ws1 = wb.active
        for fila in filas:
            ws1.append(fila)
        wb.save(filename=fn)


# pruebas básicas
if __name__ == "__main__":
    ##import pdb; pdb.set_trace()
    filas = leer("facturas-wsfev1-bis.csv")
    regs1 = desaplanar(filas)
    print(filas)
    filas1 = aplanar(regs1)
    print(filas1)
    print(filas1 == filas)
    escribir(filas1, "facturas-wsfev1-bis-sal.csv")
    escribir(filas1, "facturas-wsfev1-bis-sal.xlsx")
    filas2 = leer("facturas-wsfev1-bis-sal.xlsx")
    for fila1, fila2 in zip(filas1, filas2):
        for celda1, celda2 in zip(fila1, fila2):
            if celda1 != celda2:
                print(celda1, celda2)